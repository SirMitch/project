# File: serenity\ContextManager\context_manager.py
# Version: 1.6.8  # Fixed file creation, permissions, CSV updates, and plugin robustness
# Created: 2025-03-26
# Updated: 2025-03-31
# Description: An advanced file management system for AI-driven file editing with robust error handling and recovery.
#              Supports full and partial updates for multiple file types via extensible plugins.
#              Features include autofix, rollback, diff generation, performance profiling,
#              transaction management, multi-interface communication (CLI, GUI, API, events),
#              and robust crash recovery with crash_monitor.py integration.
# Usage:
#   python context_manager.py [project_dir] [--verbose] [--debug] [--gui] [--api] [--api-host HOST] [--api-port PORT] [--event]
#   Additional options: --list-versions, --list-locks, --reset-versions, --reset-locks, --validate-only,
#                       --rollback [file], --profile, --timeout SECONDS, --retry-attempts N, --retry-delay SECONDS
# Dependencies: psutil, jsonpatch, pyyaml, openpyxl, PyPDF2, fastapi, uvicorn, Pillow
# Notes:
#   - Use --timeout to set max execution time per operation (default: 300s).
#   - Configure retry_attempts and retry_delay via CLI or config for resilience.
#   - Crash recovery requires crash_monitor.py in the project directory.
#   - API credentials must be set via environment variables SERENITY_API_USER and SERENITY_API_PASS.
#   - strict_size_check (default: False) in config enforces max_size_mb limit with an exception if True.

import os
import sys
import subprocess
import re
import json
import time
import stat
import shutil
import logging
import logging.handlers
import argparse
import importlib
import hashlib
import base64
import zipfile
import tarfile
import xml.etree.ElementTree as ET
import sqlite3
import csv
import yaml
import jsonpatch
import psutil
import multiprocessing as mp
import cProfile
import pstats
import signal
import threading
import atexit
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple, Type
from abc import ABC, abstractmethod
from contextlib import contextmanager
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, TimeoutError as FutureTimeoutError
from fastapi import FastAPI, HTTPException, Security
from fastapi.security import HTTPBasic, HTTPBasicCredentials
import uvicorn
import difflib
from tempfile import TemporaryDirectory
from pathlib import Path

# Constants
PACKAGE_IMPORT_MAP = {
    "psutil": "psutil", "jsonpatch": "jsonpatch", "pyyaml": "yaml", "openpyxl": "openpyxl",
    "PyPDF2": "PyPDF2", "fastapi": "fastapi", "uvicorn": "uvicorn", "Pillow": "PIL"
}
DEFAULT_TIMEOUT = 300  # 5 minutes
LOCK_TIMEOUT = 60  # 1 minute

# Auto-install packages with enhanced logging
def install_package(package: str) -> None:
    import_name = PACKAGE_IMPORT_MAP.get(package, package)
    try:
        __import__(import_name)
    except ImportError:
        logging.warning(f"{package} not found. Attempting to install...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", "pip"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            subprocess.check_call([sys.executable, "-m", "pip", "install", package], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            logging.info(f"{package} installed successfully.")
            globals()[import_name] = __import__(import_name)
        except subprocess.CalledProcessError as e:
            logging.error(f"Failed to install {package}: {e}. Please install manually: pip install {package}")
            sys.exit(1)
        except Exception as e:
            logging.error(f"Unexpected error installing {package}: {e}", exc_info=True)
            sys.exit(1)

required_packages = ["psutil", "jsonpatch", "pyyaml", "openpyxl", "PyPDF2", "fastapi", "uvicorn", "Pillow"]
for package in required_packages:
    install_package(package)

# Plugin Interface
class Plugin(ABC):
    @abstractmethod
    def can_handle(self, update_type: str, file_extension: str) -> bool:
        pass

    @abstractmethod
    def apply_update(self, filepath: str, update: Dict[str, Any]) -> Dict[str, Any]:
        pass

class ExcelPlugin(Plugin):
    def can_handle(self, update_type: str, file_extension: str) -> bool:
        return update_type == "ExcelUpdate" and file_extension in (".xlsx", ".xls")

    def apply_update(self, filepath: str, update: Dict[str, Any]) -> Dict[str, Any]:
        import openpyxl
        try:
            sheet = update.get("sheet", "Sheet1")
            cell = update.get("cell", "A1")
            new_value = update.get("excel_new_value", "")
            try:
                wb = openpyxl.load_workbook(filepath)
            except openpyxl.utils.exceptions.InvalidFileException:
                wb = openpyxl.Workbook()
                wb.create_sheet(sheet)
                wb.remove(wb["Sheet"])  # Remove default sheet
                wb.save(filepath)
                wb = openpyxl.load_workbook(filepath)
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
            ws = wb[sheet]
            ws[cell].value = new_value
            wb.save(filepath)
            wb.close()
            return {"sheet": sheet, "cell": cell}
        except Exception as e:
            raise RuntimeError(f"Failed to update Excel file {filepath}: {e}")

class PDFPlugin(Plugin):
    def can_handle(self, update_type: str, file_extension: str) -> bool:
        return update_type == "PDFUpdate" and file_extension == ".pdf"

    def apply_update(self, filepath: str, update: Dict[str, Any]) -> Dict[str, Any]:
        import PyPDF2
        try:
            metadata = update.get("metadata", "Updated by ContextManager")
            with open(filepath, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                writer = PyPDF2.PdfWriter()
                for page in reader.pages:
                    writer.add_page(page)
                writer.add_metadata({"/Updated": metadata})
                with open(filepath, "wb") as f_out:
                    writer.write(f_out)
            return {"metadata": metadata}
        except PyPDF2.errors.PdfReadError as e:
            raise RuntimeError(f"Invalid PDF file {filepath}: {e}")
        except Exception as e:
            raise RuntimeError(f"Failed to update PDF file {filepath}: {e}")

class TextPlugin(Plugin):
    def can_handle(self, update_type: str, file_extension: str) -> bool:
        return update_type == "RegexUpdate" and file_extension in (".txt", ".md", ".py")

    def apply_update(self, filepath: str, update: Dict[str, Any]) -> Dict[str, Any]:
        try:
            pattern = update.get("pattern", "")
            replacement = update.get("replacement", "")
            with open(filepath, "r", encoding="utf-8") as f:
                content = f.read()
            updated_content = re.sub(pattern, replacement, content)
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(updated_content)
            return {"pattern": pattern, "replacement": replacement}
        except Exception as e:
            raise RuntimeError(f"Failed to apply regex update to {filepath}: {e}")

class ImagePlugin(Plugin):
    def can_handle(self, update_type: str, file_extension: str) -> bool:
        return update_type == "ImageUpdate" and file_extension in (".jpg", ".png", ".jpeg")

    def apply_update(self, filepath: str, update: Dict[str, Any]) -> Dict[str, Any]:
        from PIL import Image, ExifTags
        try:
            description = update.get("description", "Updated by ContextManager")
            img = Image.open(filepath)
            exif = img.getexif() or Image.Exif()
            exif[0x010e] = description  # ImageDescription
            img.save(filepath, exif=exif)
            img.close()
            return {"description": description}
        except PIL.UnidentifiedImageError as e:
            raise RuntimeError(f"Invalid image file {filepath}: {e}")
        except Exception as e:
            raise RuntimeError(f"Failed to update image metadata {filepath}: {e}")

# I/O Handlers
class IOHandler(ABC):
    @abstractmethod
    def read_session_summary(self, session_file: str) -> str:
        pass

    @abstractmethod
    def write_context_summary(self, context_file: str, content: str) -> None:
        pass

    @abstractmethod
    def report_status(self, message: str, level: str = "info") -> None:
        pass

class CommandLineIOHandler(IOHandler):
    def __init__(self, verbose: bool = False):
        self.verbose = verbose

    def read_session_summary(self, session_file: str) -> str:
        try:
            with open(session_file, "r", encoding="utf-8") as f:
                return f.read()
        except Exception as e:
            logging.error(f"Failed to read {session_file}: {e}", exc_info=True)
            raise

    def write_context_summary(self, context_file: str, content: str) -> None:
        try:
            with open(context_file, "w", encoding="utf-8") as f:
                f.write(content)
        except Exception as e:
            logging.error(f"Failed to write {context_file}: {e}", exc_info=True)
            raise

    def report_status(self, message: str, level: str = "info") -> None:
        if self.verbose or level in ("error", "warning"):
            print(f"[{level.upper()}] {message}")

class SerenityIOHandler(IOHandler):
    def __init__(self, events, logger):
        self.events = events
        self.logger = logger
        self.events.subscribe("file_management_instruction", self.handle_instruction)

    def read_session_summary(self, session_file: str) -> str:
        try:
            with open(session_file, "r", encoding="utf-8") as f:
                content = f.read()
            self.report_status(f"Read session summary from {session_file}", "info")
            return content
        except Exception as e:
            self.report_status(f"Failed to read session summary: {e}", "error")
            raise

    def write_context_summary(self, context_file: str, content: str) -> None:
        try:
            with open(context_file, "w", encoding="utf-8") as f:
                f.write(content)
            self.report_status(f"Wrote context summary to {context_file}", "info")
        except Exception as e:
            self.report_status(f"Failed to write context summary: {e}", "error")
            raise

    def report_status(self, message: str, level: str = "info") -> None:
        self.logger.log(getattr(logging, level.upper()), message)
        self.events.emit("file_management_status", {
            "message": message, "level": level, "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

    def handle_instruction(self, data: Dict[str, Any]) -> None:
        instruction = data.get("instruction")
        if instruction:
            self.report_status(f"Received instruction: {instruction}", "info")
            if hasattr(self, "manager"):
                self.manager.process_instruction(instruction)

class EventIOHandler(SerenityIOHandler):
    pass

class APIIOHandler(IOHandler):
    def __init__(self, logger, host: str = "localhost", port: int = 8000):
        self.logger = logger
        self.host = host
        self.port = port
        self.app = FastAPI()
        self.security = HTTPBasic()
        self.status_messages = []
        self.api_user = os.getenv("SERENITY_API_USER")
        self.api_pass = os.getenv("SERENITY_API_PASS")
        if not self.api_user or not self.api_pass:
            raise ValueError("SERENITY_API_USER and SERENITY_API_PASS must be set in environment variables")
        self.setup_routes()

    def setup_routes(self):
        @self.app.post("/process_instruction")
        async def process_instruction(data: Dict[str, Any], credentials: HTTPBasicCredentials = Security(self.security)):
            if credentials.username != self.api_user or credentials.password != self.api_pass:
                raise HTTPException(status_code=401, detail="Invalid credentials")
            instruction = data.get("instruction")
            if not instruction:
                raise HTTPException(status_code=400, detail="Instruction required")
            self.logger.info(f"Received API instruction: {instruction}")
            if hasattr(self, "manager"):
                success = self.manager.process_instruction(instruction)
                self.report_status(f"Processed instruction: {instruction}", "info" if success else "error")
                return {"status": "success" if success else "failed", "message": instruction}
            raise HTTPException(status_code=500, detail="ContextManager not initialized")

        @self.app.get("/status")
        async def get_status(credentials: HTTPBasicCredentials = Security(self.security)):
            if credentials.username != self.api_user or credentials.password != self.api_pass:
                raise HTTPException(status_code=401, detail="Invalid credentials")
            return {"status_messages": self.status_messages[-10:]}

    def start_server(self):
        def run_server():
            try:
                self.logger.info(f"Starting API server on {self.host}:{self.port}")
                uvicorn.run(self.app, host=self.host, port=self.port)
            except Exception as e:
                self.logger.error(f"API server failed: {e}", exc_info=True)
                self.report_status(f"API server failed: {e}", "error")
        threading.Thread(target=run_server, daemon=True).start()
        time.sleep(1)  # Wait for server to start

    def read_session_summary(self, session_file: str) -> str:
        try:
            with open(session_file, "r", encoding="utf-8") as f:
                content = f.read()
            self.report_status(f"Read session summary from {session_file}", "info")
            return content
        except Exception as e:
            self.report_status(f"Failed to read session summary: {e}", "error")
            raise

    def write_context_summary(self, context_file: str, content: str) -> None:
        try:
            with open(context_file, "w", encoding="utf-8") as f:
                f.write(content)
            self.report_status(f"Wrote context summary to {context_file}", "info")
        except Exception as e:
            self.report_status(f"Failed to write context summary: {e}", "error")
            raise

    def report_status(self, message: str, level: str = "info") -> None:
        self.logger.log(getattr(logging, level.upper()), message)
        self.status_messages.append({"message": message, "level": level, "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})

class ContextManager:
    def __init__(self, project_dir: str, io_handler: IOHandler, config: Dict[str, Any] = None):
        self.project_dir = os.path.abspath(project_dir)
        if not os.access(self.project_dir, os.W_OK):
            raise PermissionError(f"Project directory {self.project_dir} is not writable")
        os.makedirs(self.project_dir, exist_ok=True)
        self.session_file = os.path.join(self.project_dir, "session_summary.md")
        self.project_log_file = os.path.join(self.project_dir, "PROJECT_LOG.md")
        self.context_summary_file = os.path.join(self.project_dir, "context_summary.md")
        self.log_file = os.path.join(self.project_dir, "context_manager.log")
        self.lock_file = os.path.join(self.project_dir, "context_manager.lock")
        self.transaction_file = os.path.join(self.project_dir, "transaction.json")
        self.signal_file = os.path.join(self.project_dir, "context_manager_done.signal")
        self.rollback_file = os.path.join(self.project_dir, "rollback.json")
        self.versions_file = os.path.join(self.project_dir, "versions.json")
        self.locks_file = os.path.join(self.project_dir, "locks.json")
        self.locks_lock_file = os.path.join(self.project_dir, "locks.lock")
        self.temp_dir = TemporaryDirectory(dir=self.project_dir)

        self.config = {
            "autofix": True,
            "max_size_mb": 10,
            "retry_attempts": 3,
            "retry_delay": 1.0,
            "parallel_workers": min(mp.cpu_count(), 8),
            "profile_mode": False,
            "json_logging": False,
            "timeout": DEFAULT_TIMEOUT,
            "strict_size_check": False
        }
        if config:
            self.config.update(config)

        self.io_handler = io_handler
        if isinstance(self.io_handler, (SerenityIOHandler, EventIOHandler, APIIOHandler)):
            self.io_handler.manager = self
        self.setup_logging()
        self.check_crash_status()
        self.plugins = self.load_plugins()
        self.file_hashes = {}
        self._running = True
        signal.signal(signal.SIGINT, self._signal_handler)
        signal.signal(signal.SIGTERM, self._signal_handler)
        atexit.register(self._cleanup)
        self.check_script_lock()
        self.start_crash_monitor()
        self.initialize_project_log()
        self.cleanup_stale_locks()
        self.check_transaction_state()

    def _signal_handler(self, signum, frame):
        self.logger.info(f"Received signal {signum}. Shutting down...")
        self._running = False
        self._cleanup()

    def _cleanup(self):
        self.release_script_lock()
        self.temp_dir.cleanup()
        self.signal_completion()

    def setup_logging(self) -> None:
        try:
            os.makedirs(os.path.dirname(self.log_file), exist_ok=True)
            formatter = logging.Formatter(
                '{"time": "%(asctime)s", "level": "%(levelname)s", "message": "%(message)s"}'
                if self.config["json_logging"] else "%(asctime)s - %(levelname)s - %(message)s"
            )
            handler = logging.handlers.RotatingFileHandler(self.log_file, maxBytes=10*1024*1024, backupCount=5)
            handler.setFormatter(formatter)
            self.logger = logging.getLogger(__name__)
            self.logger.setLevel(logging.DEBUG)
            self.logger.handlers = []
            self.logger.addHandler(handler)
            self.logger.info("Initialized ContextManager")
        except Exception as e:
            logging.basicConfig(level=logging.DEBUG)
            logging.critical(f"Failed to set up logging: {e}", exc_info=True)
            sys.exit(1)

    def check_crash_status(self) -> None:
        crash_status_file = os.path.join(self.project_dir, "crash_status.json")
        if os.path.exists(crash_status_file):
            try:
                with open(crash_status_file, "r") as f:
                    status = json.load(f)
                self.logger.warning(f"Previous crash detected: {status['message']} at {status['timestamp']}")
                self.io_handler.report_status(f"Previous crash: {status['message']}", "warning")
                os.remove(crash_status_file)
            except json.JSONDecodeError as e:
                self.logger.error(f"Invalid JSON in crash_status_file: {e}", exc_info=True)
                if self.config["autofix"]:
                    os.remove(crash_status_file)
                    self.logger.info("Autofixed stale crash_status.json due to invalid JSON")
            except Exception as e:
                self.logger.error(f"Failed to process crash status: {e}", exc_info=True)
                if self.config["autofix"]:
                    os.remove(crash_status_file)
                    self.logger.info("Autofixed stale crash_status.json")

    def load_plugins(self) -> List[Plugin]:
        plugins_dir = os.path.join(self.project_dir, "plugins")
        os.makedirs(plugins_dir, exist_ok=True)
        sys.path.append(plugins_dir)
        plugins = [ExcelPlugin(), PDFPlugin(), TextPlugin(), ImagePlugin()]
        for plugin_file in os.listdir(plugins_dir):
            if plugin_file.endswith(".py") and plugin_file != "__init__.py":
                try:
                    module_name = plugin_file[:-3]
                    module = importlib.import_module(module_name)
                    plugin_class = getattr(module, "CustomPlugin", None)
                    if plugin_class and issubclass(plugin_class, Plugin):
                        plugins.append(plugin_class())
                        self.logger.info(f"Loaded custom plugin: {module_name}")
                except Exception as e:
                    self.logger.error(f"Failed to load plugin {plugin_file}: {e}", exc_info=True)
        return plugins

    def check_script_lock(self) -> None:
        if os.path.exists(self.lock_file):
            try:
                with open(self.lock_file, "r") as f:
                    pid = f.read().strip()
                if pid and psutil.pid_exists(int(pid)):
                    self.logger.error(f"Another instance running (PID: {pid})")
                    self.io_handler.report_status(f"Another instance running (PID: {pid})", "error")
                    sys.exit(1)
                else:
                    self.logger.warning(f"Removing stale lock file (PID: {pid})")
                    os.remove(self.lock_file)
            except (ValueError, OSError) as e:
                self.logger.warning(f"Invalid or inaccessible lock file: {e}. Removing.", exc_info=True)
                try:
                    os.remove(self.lock_file)
                except Exception as e:
                    self.logger.error(f"Failed to remove lock file: {e}", exc_info=True)
        try:
            with open(self.lock_file, "w") as f:
                f.write(str(os.getpid()))
            self.logger.info("Acquired script lock")
        except PermissionError as e:
            self.logger.error(f"Permission denied when writing lock file {self.lock_file}: {e}")
            raise

    def start_crash_monitor(self) -> None:
        monitor_script = os.path.join(self.project_dir, "crash_monitor.py")
        if not os.path.exists(monitor_script):
            self.logger.warning("crash_monitor.py not found. Crash monitoring disabled.")
            return
        cmd = [sys.executable, monitor_script, str(os.getpid()), self.project_dir, "--timeout", str(self.config["timeout"] * 10)]
        try:
            subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=False)
            self.logger.info("Started crash_monitor.py")
        except Exception as e:
            self.logger.error(f"Failed to start crash_monitor.py: {e}", exc_info=True)

    def signal_completion(self) -> None:
        for attempt in range(self.config["retry_attempts"]):
            try:
                with open(self.signal_file, "w") as f:
                    f.write(f"done:{datetime.now().isoformat()}")
                self.logger.info("Signaled completion")
                return
            except Exception as e:
                self.logger.error(f"Attempt {attempt + 1}/{self.config['retry_attempts']}: Failed to signal completion: {e}", exc_info=True)
                if attempt == self.config["retry_attempts"] - 1:
                    raise RuntimeError("Failed to signal completion after retries")
                time.sleep(self.config["retry_delay"])

    def release_script_lock(self) -> None:
        if os.path.exists(self.lock_file):
            for attempt in range(self.config["retry_attempts"]):
                try:
                    os.remove(self.lock_file)
                    self.logger.info("Released script lock")
                    break
                except Exception as e:
                    self.logger.error(f"Attempt {attempt + 1}: Failed to release script lock: {e}", exc_info=True)
                    if attempt == self.config["retry_attempts"] - 1:
                        self.logger.critical("Script lock may persist")
                    time.sleep(self.config["retry_delay"])

    def initialize_project_log(self) -> None:
        if not os.path.exists(self.project_log_file):
            initial_content = """# Project Log\n\n## Memory Log\n\n### Table of Contents\n### Entry\n\n## Memory Log Backup\n\n### Table of Contents\n### Entry\n\n## Changelog\n\n### [Unreleased]\n#### Added\n#### Changed\n#### Fixed\n"""
            try:
                with open(self.project_log_file, "w", encoding="utf-8") as f:
                    f.write(initial_content)
                self.logger.info(f"Initialized {self.project_log_file}")
            except Exception as e:
                self.logger.error(f"Failed to initialize project log: {e}", exc_info=True)
                raise

    def cleanup_stale_locks(self) -> None:
        if not os.path.exists(self.locks_file):
            with open(self.locks_file, "w") as f:
                json.dump({}, f)
            return
        try:
            with self.file_lock_context(self.locks_file):
                with open(self.locks_file, "r") as f:
                    locks = json.load(f)
                current_time = datetime.now().timestamp()
                updated_locks = {}
                for filename, timestamp in locks.items():
                    lock_time = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S").timestamp()
                    if current_time - lock_time < 86400:  # 24 hours
                        lock_pid_file = os.path.join(self.project_dir, f"{filename}.lock.pid")
                        if os.path.exists(lock_pid_file):
                            with open(lock_pid_file, "r") as pf:
                                pid = pf.read().strip()
                            if pid and not psutil.pid_exists(int(pid)):
                                self.logger.info(f"Removing stale lock for {filename} (PID {pid} no longer exists)")
                                continue
                        updated_locks[filename] = timestamp
                    else:
                        self.logger.info(f"Removed stale lock for {filename} (locked at {timestamp})")
                with open(self.locks_file, "w") as f:
                    json.dump(updated_locks, f)
            self.logger.info("Cleaned up stale locks")
        except Exception as e:
            self.logger.error(f"Failed to clean up stale locks: {e}", exc_info=True)

    def load_transaction(self) -> Dict[str, Any]:
        if not os.path.exists(self.transaction_file):
            self.logger.info(f"No transaction file found at {self.transaction_file}. Using default state.")
            return {"state": "idle", "backups": {}, "timestamp": datetime.now().isoformat()}
        for attempt in range(self.config["retry_attempts"]):
            try:
                with open(self.transaction_file, "r") as f:
                    return json.load(f)
            except Exception as e:
                self.logger.error(f"Attempt {attempt + 1}/{self.config['retry_attempts']}: Failed to load transaction: {e}", exc_info=True)
                if attempt == self.config["retry_attempts"] - 1 and self.config["autofix"]:
                    default_state = {"state": "idle", "backups": {}, "timestamp": datetime.now().isoformat()}
                    with open(self.transaction_file, "w") as f:
                        json.dump(default_state, f, indent=4)
                    self.logger.info("Autofixed corrupted transaction.json")
                    return default_state
                time.sleep(self.config["retry_delay"])
        raise RuntimeError("Failed to load transaction.json after retries")

    def save_transaction(self, transaction: Dict[str, Any]) -> None:
        transaction["timestamp"] = datetime.now().isoformat()
        for attempt in range(self.config["retry_attempts"]):
            try:
                with open(self.transaction_file, "w") as f:
                    json.dump(transaction, f, indent=4)
                self.logger.info("Saved transaction state")
                return
            except Exception as e:
                self.logger.error(f"Attempt {attempt + 1}/{self.config['retry_attempts']}: Failed to save transaction: {e}", exc_info=True)
                if attempt == self.config["retry_attempts"] - 1:
                    raise RuntimeError(f"Failed to save transaction after {self.config['retry_attempts']} attempts")
                time.sleep(self.config["retry_delay"])

    def restore_file_if_needed(self, filepath: str, backup_path: str) -> bool:
        if not os.path.exists(backup_path):
            self.logger.warning(f"Backup {backup_path} not found for {filepath}")
            return False
        if not os.path.exists(filepath) or abs(os.path.getmtime(filepath) - os.path.getmtime(backup_path)) > 1:
            shutil.copyfile(backup_path, filepath)
            self.logger.info(f"Restored {filepath} from {backup_path}")
            return True
        return False

    def check_transaction_state(self) -> None:
        transaction = self.load_transaction()
        if transaction["state"] != "idle":
            self.logger.warning("Incomplete previous run detected. Rolling back.")
            with ThreadPoolExecutor(max_workers=self.config["parallel_workers"]) as executor:
                futures = [executor.submit(self.restore_file_if_needed, filepath, backup_path) for filepath, backup_path in transaction["backups"].items()]
                for future in futures:
                    try:
                        future.result(timeout=self.config["timeout"])
                    except FutureTimeoutError:
                        self.logger.error(f"Rollback timed out after {self.config['timeout']} seconds")
            with self.file_lock_context(self.locks_file):
                with open(self.locks_file, "w") as f:
                    json.dump({}, f)
            transaction["state"] = "idle"
            transaction["backups"] = {}
            self.save_transaction(transaction)
            self.logger.info("Transaction state reset after rollback")

    def compute_file_hash(self, filepath: str) -> str:
        if filepath in self.file_hashes:
            return self.file_hashes[filepath]
        hasher = hashlib.sha256()
        try:
            with open(filepath, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hasher.update(chunk)
        except Exception as e:
            self.logger.error(f"Failed to compute hash for {filepath}: {e}", exc_info=True)
            raise
        hash_value = hasher.hexdigest()
        self.file_hashes[filepath] = hash_value
        return hash_value

    def backup_file(self, filepath: str) -> Tuple[str, float]:
        if not os.path.exists(filepath):
            if self.config["autofix"]:
                os.makedirs(os.path.dirname(filepath), exist_ok=True)
                with open(filepath, "wb") as f:
                    pass
                self.logger.info(f"Created empty file {filepath} for backup")
            else:
                raise FileNotFoundError(f"{filepath} does not exist")
        
        current_hash = self.compute_file_hash(filepath)
        backup_dir = os.path.join(self.project_dir, "backups")
        os.makedirs(backup_dir, exist_ok=True)
        for backup_file in os.listdir(backup_dir):
            if backup_file.startswith(os.path.basename(filepath)) and backup_file.endswith(".backup"):
                backup_path = os.path.join(backup_dir, backup_file)
                if current_hash == self.compute_file_hash(backup_path):
                    return backup_path, os.path.getmtime(filepath)
        
        backup_path = os.path.join(backup_dir, f"{os.path.basename(filepath)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.backup")
        shutil.copyfile(filepath, backup_path)
        self.file_hashes[filepath] = current_hash
        self.logger.info(f"Backed up {filepath} to {backup_path}")
        return backup_path, os.path.getmtime(filepath)

    def restore_file(self, filepath: str, backup_path: str) -> None:
        shutil.copyfile(backup_path, filepath)
        self.file_hashes.pop(filepath, None)
        self.logger.info(f"Restored {filepath} from {backup_path}")

    def check_file_modified(self, filepath: str, original_mtime: float) -> None:
        current_mtime = os.path.getmtime(filepath)
        if abs(current_mtime - original_mtime) > 1:
            self.logger.error(f"File {filepath} modified externally (mtime: {original_mtime} -> {current_mtime})")
            raise RuntimeError(f"File {filepath} modified externally")

    def apply_update(self, update: Dict[str, Any], backup_info: Dict[str, Tuple[str, float]]) -> Tuple[str, str, Dict[str, Any], Optional[str]]:
        filepath = self.resolve_filepath(update["file"], update.get("directory"))
        if not os.path.exists(filepath):
            if update["if_not_exists"] == "Create":
                os.makedirs(os.path.dirname(filepath), exist_ok=True)
                with open(filepath, "wb") as f:
                    if "content" in update:
                        if update["content_type"] == "Binary":
                            f.write(base64.b64decode(update["content"]))
                        else:
                            f.write(update["content"].encode("utf-8"))
                    else:
                        f.write(b"")
                backup_info[filepath] = self.backup_file(filepath)
                self.logger.info(f"Created and backed up {filepath}")
            elif update["if_not_exists"] == "Skip":
                self.logger.info(f"Skipping update for non-existent file {filepath}")
                return filepath, None, {}, None
            else:
                raise FileNotFoundError(f"{filepath} does not exist")
        
        self.check_file_permissions(filepath)
        current_version = self.check_version(filepath, update.get("current_version", "0.0.0"))
        with self.file_lock(filepath):
            self.check_file_modified(filepath, backup_info[filepath][1])
            update_type = update.get("update_type", "Full")
            update_details = {}
            diff = None
            file_extension = os.path.splitext(filepath)[1].lower()
            plugin = next((p for p in self.plugins if p.can_handle(update_type, file_extension)), None)

            def apply_with_timeout():
                nonlocal diff, update_details
                try:
                    if plugin:
                        update_details = plugin.apply_update(filepath, update)
                    elif update_type == "Full":
                        with open(filepath, "r", encoding="utf-8") as f:
                            old_content = f.read()
                        if update["content_type"] == "Text":
                            with open(filepath, "w", encoding="utf-8") as f:
                                f.write(update["content"])
                            diff = self.generate_diff(old_content, update["content"])
                        elif update["content_type"] == "Binary":
                            with open(filepath, "wb") as f:
                                f.write(base64.b64decode(update["content"]))
                    elif update_type == "Method":
                        with open(filepath, "r", encoding="utf-8") as f:
                            content = f.read()
                        pattern = rf"def\s+{update['method_name']}\s*\(self[^)]*\)\s*:\s*[\s\S]*?(?=\n\s*def\s|\n\s*class\s|\Z)"
                        match = re.search(pattern, content)
                        if not match:
                            self.logger.error(f"Method {update['method_name']} not found in {filepath}")
                            raise ValueError(f"Method {update['method_name']} not found in {filepath}")
                        updated_content = content[:match.start()] + update["content"] + content[match.end():]
                        with open(filepath, "w", encoding="utf-8") as f:
                            f.write(updated_content)
                        diff = self.generate_diff(content, updated_content)
                        update_details = {"method_name": update["method_name"]}
                    elif update_type == "JSONPatch":
                        with open(filepath, "r") as f:
                            data = json.load(f)
                        patch = jsonpatch.JsonPatch.from_string(update["json_patch"])
                        updated_data = patch.apply(data)
                        with open(filepath, "w") as f:
                            json.dump(updated_data, f, indent=4)
                    elif update_type == "XMLUpdate":
                        tree = ET.parse(filepath)
                        node = tree.getroot().find(update["xpath"])
                        if node is None:
                            raise ValueError(f"XPath {update['xpath']} not found in {filepath}")
                        node.text = update["xml_new_value"]
                        tree.write(filepath)
                        update_details = {"xpath": update["xpath"]}
                    elif update_type == "YAMLUpdate":
                        with open(filepath, "r") as f:
                            data = yaml.safe_load(f) or {}
                        d = data
                        for key in update["yaml_key"].split(".")[:-1]:
                            d = d.setdefault(key, {})
                        d[update["yaml_key"].split(".")[-1]] = update["yaml_new_value"]
                        with open(filepath, "w") as f:
                            yaml.dump(data, f)
                        update_details = {"key": update["yaml_key"]}
                    elif update_type == "CSVUpdate":
                        with open(filepath, "r", encoding="utf-8") as f:
                            rows = list(csv.reader(f))
                        row_index = int(update["csv_row"])
                        new_row = update["csv_new_row"].split(",")
                        if not rows:
                            rows.append(new_row)
                        else:
                            while len(rows) <= row_index:
                                rows.append([""] * len(rows[0]))
                            rows[row_index] = new_row
                        with open(filepath, "w", newline="", encoding="utf-8") as f:
                            csv.writer(f).writerows(rows)
                        update_details = {"row": update["csv_row"]}
                    elif update_type == "ArchiveUpdate":
                        with TemporaryDirectory(dir=self.project_dir) as temp_dir:
                            if filepath.endswith(".zip"):
                                with zipfile.ZipFile(filepath, "r") as z:
                                    z.extractall(temp_dir)
                                inner_path = os.path.join(temp_dir, update["inner_path"])
                                with open(inner_path, "w", encoding="utf-8") as f:
                                    f.write(update["content"])
                                with zipfile.ZipFile(filepath, "w") as z:
                                    for root, _, files in os.walk(temp_dir):
                                        for file in files:
                                            z.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), temp_dir))
                            elif filepath.endswith(".tar"):
                                with tarfile.open(filepath, "r") as t:
                                    t.extractall(temp_dir)
                                inner_path = os.path.join(temp_dir, update["inner_path"])
                                with open(inner_path, "w", encoding="utf-8") as f:
                                    f.write(update["content"])
                                with tarfile.open(filepath, "w") as t:
                                    for root, _, files in os.walk(temp_dir):
                                        for file in files:
                                            t.add(os.path.join(root, file), os.path.relpath(os.path.join(root, file), temp_dir))
                        update_details = {"inner_path": update["inner_path"]}
                    elif update_type == "SQLCommand":
                        conn = sqlite3.connect(filepath)
                        conn.cursor().execute(update["sql_command"])
                        conn.commit()
                        conn.close()
                    else:
                        raise ValueError(f"Unsupported update type: {update_type}")
                except Exception as e:
                    self.logger.error(f"Failed to apply update to {filepath}: {e}", exc_info=True)
                    raise

            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(apply_with_timeout)
                try:
                    future.result(timeout=self.config["timeout"])
                except FutureTimeoutError:
                    self.logger.error(f"Update to {filepath} timed out after {self.config['timeout']} seconds")
                    raise TimeoutError(f"Update to {filepath} exceeded timeout")

            new_version = self.increment_version(filepath, current_version)
            self.update_changelog_for_version(filepath, current_version, new_version, update_type, update_details, backup_info, diff)
            return filepath, new_version, update_details, diff

    def update_files(self, updates: Dict[str, Any]) -> List[Dict[str, Any]]:
        transaction = self.load_transaction()
        transaction["state"] = "started"
        transaction["backups"] = {}
        self.save_transaction(transaction)

        backup_info = {}
        rollback_info = {"session": []}
        with ThreadPoolExecutor(max_workers=self.config["parallel_workers"]) as executor:
            backup_futures = {}
            if os.path.exists(self.versions_file):
                backup_futures[executor.submit(self.backup_file, self.versions_file)] = self.versions_file
            if os.path.exists(self.project_log_file):
                backup_futures[executor.submit(self.backup_file, self.project_log_file)] = self.project_log_file
            for update in updates["code"]:
                filepath = self.resolve_filepath(update["file"], update.get("directory"))
                if os.path.exists(filepath):
                    future = executor.submit(self.backup_file, filepath)
                    backup_futures[future] = filepath
            
            for future in backup_futures:
                try:
                    backup_path, mtime = future.result(timeout=self.config["timeout"])
                    filepath = backup_futures[future]
                    backup_info[filepath] = (backup_path, mtime)
                    transaction["backups"][filepath] = backup_path
                except FutureTimeoutError:
                    self.logger.error(f"Backup operation for {backup_futures[future]} timed out after {self.config['timeout']} seconds")
                    raise
        self.save_transaction(transaction)

        try:
            transaction["state"] = "code_updates"
            self.save_transaction(transaction)
            update_results = []
            with ProcessPoolExecutor(max_workers=self.config["parallel_workers"]) as executor:
                futures = [executor.submit(self.apply_update, update, backup_info) for update in updates["code"]]
                for future in futures:
                    try:
                        filepath, new_version, update_details, diff = future.result(timeout=self.config["timeout"])
                        if new_version:
                            result = {"filepath": filepath, "new_version": new_version, "details": update_details}
                            if diff:
                                result["diff"] = diff
                            update_results.append(result)
                            if filepath in backup_info:
                                rollback_info["session"].append({"filepath": filepath, "backup_path": backup_info[filepath][0]})
                            else:
                                self.logger.warning(f"No backup found for {filepath} in rollback info")
                    except FutureTimeoutError:
                        self.logger.error(f"Update operation timed out after {self.config['timeout']} seconds")
                        raise

            if updates["changelog"]:
                transaction["state"] = "changelog_updates"
                self.save_transaction(transaction)
                if self.project_log_file not in backup_info:
                    backup_info[self.project_log_file] = self.backup_file(self.project_log_file)
                    transaction["backups"][self.project_log_file] = backup_info[self.project_log_file][0]
                with self.file_lock(self.project_log_file):
                    self.check_file_permissions(self.project_log_file)
                    self.check_file_modified(self.project_log_file, backup_info[self.project_log_file][1])
                    with open(self.project_log_file, "r", encoding="utf-8") as f:
                        content = f.read()
                    changelog_section = content.split("## Changelog")[1]
                    updated_content = content.replace(changelog_section, updates["changelog"] + "\n\n" + changelog_section.lstrip())
                    with open(self.project_log_file, "w", encoding="utf-8") as f:
                        f.write(updated_content)
                    self.increment_version(self.project_log_file, self.load_versions().get("PROJECT_LOG.md", "0.0.0"))
                    rollback_info["session"].append({"filepath": self.project_log_file, "backup_path": backup_info[self.project_log_file][0]})

            if updates["memory_log"]:
                transaction["state"] = "memory_log_updates"
                self.save_transaction(transaction)
                if self.project_log_file not in backup_info:
                    backup_info[self.project_log_file] = self.backup_file(self.project_log_file)
                    transaction["backups"][self.project_log_file] = backup_info[self.project_log_file][0]
                with self.file_lock(self.project_log_file):
                    self.check_file_permissions(self.project_log_file)
                    self.check_file_modified(self.project_log_file, backup_info[self.project_log_file][1])
                    with open(self.project_log_file, "r", encoding="utf-8") as f:
                        content = f.read()
                    memory_log_section = content.split("## Memory Log")[1].split("## Memory Log Backup")[0]
                    toc = memory_log_section.split("### Table of Contents")[1].split("### Entry")[0]
                    new_toc_entry = f"- [Entry {updates['message_ref'].split('#')[1]}: {updates['memory_log'].split('## Entry')[1].split('\n')[0].strip()}](#entry-{updates['message_ref'].split('#')[1]}-{updates['memory_log'].split('## Entry')[1].split('\n')[0].strip().lower().replace(' ', '-')})\n"
                    updated_toc = toc.rstrip() + new_toc_entry + "### Entry"
                    updated_memory_log = memory_log_section.replace(toc, updated_toc).rstrip() + "\n\n" + updates["memory_log"]
                    updated_content = content.replace(memory_log_section, updated_memory_log)
                    memory_log_backup_section = updated_content.split("## Memory Log Backup")[1].split("## Changelog")[0]
                    updated_content = updated_content.replace(memory_log_backup_section, updated_memory_log)
                    with open(self.project_log_file, "w", encoding="utf-8") as f:
                        f.write(updated_content)
                    self.increment_version(self.project_log_file, self.load_versions().get("PROJECT_LOG.md", "0.0.0"))
                    rollback_info["session"].append({"filepath": self.project_log_file, "backup_path": backup_info[self.project_log_file][0]})

            transaction["state"] = "idle"
            transaction["backups"] = {}
            self.save_transaction(transaction)
            with open(self.rollback_file, "w") as f:
                json.dump(rollback_info, f, indent=4)
            return update_results
        except Exception as e:
            self.logger.error(f"Update failed: {e}", exc_info=True)
            with ThreadPoolExecutor(max_workers=self.config["parallel_workers"]) as executor:
                futures = [executor.submit(self.restore_file_if_needed, filepath, backup_path) for filepath, (backup_path, _) in backup_info.items()]
                for future in futures:
                    try:
                        future.result(timeout=self.config["timeout"])
                    except FutureTimeoutError:
                        self.logger.error(f"Rollback timed out after {self.config['timeout']} seconds")
            transaction["state"] = "failed"
            self.save_transaction(transaction)
            raise

    def process_session_summary(self) -> bool:
        if not os.path.exists(self.session_file):
            self.logger.error(f"Session summary not found at {self.session_file}")
            return False
        content = self.io_handler.read_session_summary(self.session_file)
        is_valid, errors = self.validate_session_summary(content)
        if not is_valid:
            for error in errors:
                self.logger.error(error)
            return False
        if not self.validate_files():
            return False
        updates = self.parse_session_summary(content)
        
        def process_with_timeout():
            if self.config["profile_mode"]:
                profiler = cProfile.Profile()
                profiler.enable()
                update_results = self.update_files(updates)
                profiler.disable()
                ps = pstats.Stats(profiler).sort_stats("cumulative")
                ps.print_stats()
                self.logger.info("Profiled session processing")
            else:
                start_time = time.time()
                update_results = self.update_files(updates)
                elapsed_time = time.time() - start_time
                self.logger.info(f"Processed session summary in {elapsed_time:.2f} seconds")
            self.generate_context_summary(updates, update_results, "Success")
            self.signal_completion()

        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(process_with_timeout)
            try:
                future.result(timeout=self.config["timeout"])
                return True
            except FutureTimeoutError:
                self.logger.error(f"Session processing timed out after {self.config['timeout']} seconds")
                return False
            except Exception as e:
                self.logger.error(f"Session processing failed: {e}", exc_info=True)
                return False

    def load_versions(self) -> Dict[str, str]:
        if not os.path.exists(self.versions_file):
            return {}
        try:
            with open(self.versions_file, "r") as f:
                versions = json.load(f)
            for filename, version in versions.items():
                if not re.match(r"^\d+\.\d+\.\d+$", version):
                    self.logger.error(f"Invalid version format for {filename}: {version}")
                    raise ValueError(f"Invalid version format for {filename}: {version}")
            return versions
        except Exception as e:
            self.logger.error(f"Failed to load versions: {e}", exc_info=True)
            raise

    def save_versions(self, versions: Dict[str, str]) -> None:
        try:
            with open(self.versions_file, "w") as f:
                json.dump(versions, f, indent=4)
            self.logger.info("Updated versions.json")
        except Exception as e:
            self.logger.error(f"Failed to save versions: {e}", exc_info=True)
            raise

    def check_version(self, filepath: str, expected_version: str) -> str:
        if not re.match(r"^\d+\.\d+\.\d+$", expected_version):
            self.logger.error(f"Invalid version format: {expected_version}")
            raise ValueError(f"Invalid version format: {expected_version}")
        versions = self.load_versions()
        filename = os.path.basename(filepath)
        current_version = versions.get(filename, "0.0.0")
        if current_version == "0.0.0" and expected_version != "0.0.0":
            versions[filename] = expected_version
            self.save_versions(versions)
            return expected_version
        if current_version != expected_version:
            if tuple(map(int, current_version.split("."))) > tuple(map(int, expected_version.split("."))):
                self.logger.warning(f"{filepath} newer than expected ({current_version} > {expected_version})")
                return current_version
            self.logger.error(f"Version mismatch for {filepath}: expected {expected_version}, found {current_version}")
            raise ValueError(f"Version mismatch for {filepath}: expected {expected_version}, found {current_version}")
        return current_version

    def increment_version(self, filepath: str, current_version: str) -> str:
        major, minor, patch = map(int, current_version.split("."))
        new_version = f"{major}.{minor}.{patch + 1}"
        versions = self.load_versions()
        versions[os.path.basename(filepath)] = new_version
        self.save_versions(versions)
        return new_version

    def load_locks(self) -> Dict[str, str]:
        with self.file_lock_context(self.locks_file):
            if not os.path.exists(self.locks_file):
                return {}
            with open(self.locks_file, "r") as f:
                return json.load(f)

    def save_locks(self, locks: Dict[str, str]) -> None:
        with self.file_lock_context(self.locks_file):
            with open(self.locks_file, "w") as f:
                json.dump(locks, f, indent=4)
            self.logger.info("Updated locks.json")

    @contextmanager
    def file_lock_context(self, filepath: str):
        lock_file = self.locks_lock_file
        fd = None
        start_time = time.time()
        while time.time() - start_time < LOCK_TIMEOUT:
            try:
                fd = os.open(lock_file, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                self.logger.debug(f"Acquired lock for {filepath}")
                yield
                break
            except FileExistsError:
                self.logger.debug(f"Waiting for lock on {filepath}")
                time.sleep(self.config["retry_delay"])
            except PermissionError as e:
                self.logger.error(f"Permission denied acquiring lock for {filepath}: {e}")
                raise
            except Exception as e:
                self.logger.error(f"Failed to acquire lock for {filepath}: {e}", exc_info=True)
                raise
        else:
            if fd is not None:
                os.close(fd)
                try:
                    os.remove(lock_file)
                except Exception:
                    self.logger.warning(f"Failed to clean up lock file {lock_file} after timeout", exc_info=True)
            self.logger.error(f"Could not acquire lock for {filepath} after {LOCK_TIMEOUT} seconds")
            raise RuntimeError(f"Failed to lock {filepath} after timeout")
        if fd is not None:
            try:
                os.close(fd)
                for attempt in range(self.config["retry_attempts"]):
                    try:
                        os.remove(lock_file)
                        self.logger.debug(f"Released lock for {filepath}")
                        break
                    except Exception as e:
                        self.logger.error(f"Attempt {attempt + 1}: Failed to release lock for {filepath}: {e}", exc_info=True)
                        if attempt == self.config["retry_attempts"] - 1:
                            self.logger.critical(f"Lock file {lock_file} may persist")
                        time.sleep(self.config["retry_delay"])
            except Exception as e:
                self.logger.error(f"Failed to close or remove lock file: {e}", exc_info=True)

    @contextmanager
    def file_lock(self, filepath: str):
        self.check_out(filepath)
        try:
            yield
        finally:
            self.check_in(filepath)

    def check_out(self, filepath: str) -> None:
        for attempt in range(self.config["retry_attempts"]):
            try:
                locks = self.load_locks()
                filename = os.path.basename(filepath)
                if filename in locks:
                    self.logger.error(f"File {filepath} already locked")
                    raise ValueError(f"File {filepath} already locked")
                locks[filename] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                lock_pid_file = os.path.join(self.project_dir, f"{filename}.lock.pid")
                with open(lock_pid_file, "w") as pf:
                    pf.write(str(os.getpid()))
                self.save_locks(locks)
                self.logger.info(f"Locked {filepath}")
                return
            except ValueError:
                if attempt == self.config["retry_attempts"] - 1:
                    raise
                self.logger.info(f"Retrying lock on {filepath} (attempt {attempt + 1}/{self.config['retry_attempts']})")
                time.sleep(self.config["retry_delay"])
            except PermissionError as e:
                self.logger.error(f"Permission denied locking {filepath}: {e}")
                raise
            except Exception as e:
                self.logger.error(f"Failed to lock {filepath}: {e}", exc_info=True)
                raise

    def check_in(self, filepath: str) -> None:
        locks = self.load_locks()
        filename = os.path.basename(filepath)
        if filename not in locks:
            self.logger.error(f"File {filepath} not locked")
            raise ValueError(f"File {filepath} not locked")
        del locks[filename]
        lock_pid_file = os.path.join(self.project_dir, f"{filename}.lock.pid")
        if os.path.exists(lock_pid_file):
            try:
                os.remove(lock_pid_file)
            except Exception as e:
                self.logger.warning(f"Failed to remove lock PID file {lock_pid_file}: {e}", exc_info=True)
        self.save_locks(locks)
        self.logger.info(f"Unlocked {filepath}")

    def check_file_permissions(self, filepath: str) -> None:
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"{filepath} does not exist")
        if not os.access(filepath, os.W_OK):
            if self.config["autofix"]:
                os.chmod(filepath, stat.S_IRUSR | stat.S_IWUSR)
                self.logger.info(f"Fixed permissions for {filepath}")
            else:
                self.logger.error(f"No write permission for {filepath}")
                raise PermissionError(f"No write permission for {filepath}")

    def check_file_size(self, filepath: str) -> None:
        size_mb = os.path.getsize(filepath) / (1024 * 1024)
        if size_mb > self.config["max_size_mb"]:
            self.logger.warning(f"File {filepath} exceeds {self.config['max_size_mb']} MB ({size_mb:.2f} MB)")
            if self.config["strict_size_check"]:
                raise RuntimeError(f"File {filepath} exceeds size limit of {self.config['max_size_mb']} MB")

    def generate_diff(self, old_content: str, new_content: str) -> str:
        diff = difflib.unified_diff(
            old_content.splitlines(keepends=True),
            new_content.splitlines(keepends=True),
            fromfile="before",
            tofile="after"
        )
        return "".join(diff)

    def update_changelog_for_version(self, filepath: str, old_version: str, new_version: str, update_type: str, update_details: Dict[str, Any], backup_info: Dict[str, Tuple[str, float]], diff: Optional[str] = None) -> None:
        filename = os.path.basename(filepath)
        details = ""
        if update_type == "Method":
            details = f" (method {update_details.get('method_name', 'unknown')})"
        elif update_type == "JSONPatch":
            details = " (applied JSON patch)"
        elif update_type == "XMLUpdate":
            details = f" (updated XML node at {update_details.get('xpath', 'unknown')})"
        elif update_type == "YAMLUpdate":
            details = f" (updated YAML key {update_details.get('key', 'unknown')})"
        elif update_type == "CSVUpdate":
            details = f" (updated CSV row {update_details.get('row', 'unknown')})"
        elif update_type == "ArchiveUpdate":
            details = f" (updated {update_details.get('inner_path', 'unknown')} in archive)"
        elif update_type == "SQLCommand":
            details = " (executed SQL command)"
        elif update_type == "ExcelUpdate":
            details = f" (updated cell {update_details.get('cell', 'unknown')} in sheet {update_details.get('sheet', 'unknown')})"
        elif update_type == "PDFUpdate":
            details = " (updated metadata)"
        elif update_type == "RegexUpdate":
            details = f" (applied regex {update_details.get('pattern', 'unknown')})"
        elif update_type == "ImageUpdate":
            details = " (updated image metadata)"
        
        diff_section = f"\n#### Diff\n```diff\n{diff}\n```" if diff else ""
        entry = f"""## {datetime.now().strftime("%Y-%m-%d")}
### Changed
- **`{filename}` ({old_version} to {new_version})**
  - Updated file via {update_type} update{details}.{diff_section}
"""
        if self.project_log_file not in backup_info:
            backup_info[self.project_log_file] = self.backup_file(self.project_log_file)
        with self.file_lock(self.project_log_file):
            self.check_file_permissions(self.project_log_file)
            self.check_file_size(self.project_log_file)
            self.check_file_modified(self.project_log_file, backup_info[self.project_log_file][1])
            with open(self.project_log_file, "r", encoding="utf-8") as f:
                content = f.read()
            changelog_section = content.split("## Changelog")[1]
            updated_changelog = entry + "\n\n" + changelog_section.lstrip()
            updated_content = content.replace(changelog_section, updated_changelog)
            with open(self.project_log_file, "w", encoding="utf-8") as f:
                f.write(updated_content)
            self.logger.info(f"Updated changelog for {filename}: {old_version} to {new_version}")
            self.increment_version(self.project_log_file, self.load_versions().get("PROJECT_LOG.md", "0.0.0"))

    def validate_session_summary(self, content: str) -> Tuple[bool, List[str]]:
        required_sections = ["Metadata", "Code Updates", "Changelog Updates", "Memory Log Updates", "Action Items"]
        seen_sections = set()
        errors = []
        sections = re.split(r"##\s", content)
        if not sections[0].startswith("# Session Summary"):
            errors.append("Missing '# Session Summary' header")
        for section in sections[1:]:
            if not section.strip():
                continue
            title = section.split("\n", 1)[0].strip()
            if title in seen_sections:
                errors.append(f"Duplicate section: {title}")
            seen_sections.add(title)
        for section in required_sections:
            if f"## {section}" not in content:
                errors.append(f"Missing section: {section}")
        if not re.search(r"- \*\*Timestamp\*\*: \d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}", content):
            errors.append("Invalid or missing timestamp in Metadata")
        return not errors, errors

    def parse_session_summary(self, content: str) -> Dict[str, Any]:
        updates = {"code": [], "changelog": None, "memory_log": None, "action_items": [], "version": "Unknown", "schedule": None}
        sections = re.split(r"##\s", content)[1:]
        for section in sections:
            title, body = section.split("\n", 1)
            title = title.strip()
            if title == "Metadata":
                timestamp_match = re.search(r"- \*\*Timestamp\*\*: (.*)", body)
                message_ref_match = re.search(r"- \*\*Message Reference\*\*: (.*)", body)
                version_match = re.search(r"- \*\*Version\*\*: (.*)", body)
                schedule_match = re.search(r"- \*\*Schedule\*\*: (.*)", body)
                updates["timestamp"] = timestamp_match.group(1) if timestamp_match else "Unknown"
                updates["message_ref"] = message_ref_match.group(1) if message_ref_match else "Unknown"
                updates["version"] = version_match.group(1) if version_match else "Unknown"
                updates["schedule"] = schedule_match.group(1) if schedule_match else None
            elif title == "Code Updates" and "No updates" not in body:
                code_blocks = re.finditer(
                    r"- \*\*File\*\*: (.*?)\n"
                    r"- \*\*Current Version\*\*: (.*?)\n"
                    r"- \*\*Update Type\*\*: (.*?)\n"
                    r"(?:- \*\*Method Name\*\*: (.*?)\n)?(?:- \*\*JSON Patch\*\*: (.*?)\n)?(?:- \*\*XPath\*\*: (.*?)\n- \*\*New Value\*\*: (.*?)\n)?(?:- \*\*YAML Key\*\*: (.*?)\n- \*\*New Value\*\*: (.*?)\n)?(?:- \*\*CSV Row\*\*: (.*?)\n- \*\*New Row\*\*: (.*?)\n)?(?:- \*\*Inner Path\*\*: (.*?)\n- \*\*Inner Update Type\*\*: (.*?)\n)?(?:- \*\*SQL Command\*\*: (.*?)\n)?(?:- \*\*Sheet\*\*: (.*?)\n- \*\*Cell\*\*: (.*?)\n- \*\*New Value\*\*: (.*?)\n)?(?:- \*\*Metadata\*\*: (.*?)\n)?(?:- \*\*Pattern\*\*: (.*?)\n- \*\*Replacement\*\*: (.*?)\n)?(?:- \*\*Description\*\*: (.*?)\n)?"
                    r"- \*\*Content Type\*\*: (.*?)\n"
                    r"- \*\*If Not Exists\*\*: (.*?)(?:\n- \*\*Directory\*\*: (.*?))?\n"
                    r"- \*\*Content\*\*:\n---\n(.*?)---",
                    body, re.DOTALL
                )
                for block in code_blocks:
                    update_entry = {
                        "file": block.group(1).strip(),
                        "current_version": block.group(2).strip(),
                        "update_type": block.group(3).strip(),
                        "method_name": block.group(4).strip() if block.group(4) else None,
                        "json_patch": block.group(5).strip() if block.group(5) else None,
                        "xpath": block.group(6).strip() if block.group(6) else None,
                        "xml_new_value": block.group(7).strip() if block.group(7) else None,
                        "yaml_key": block.group(8).strip() if block.group(8) else None,
                        "yaml_new_value": block.group(9).strip() if block.group(9) else None,
                        "csv_row": block.group(10).strip() if block.group(10) else None,
                        "csv_new_row": block.group(11).strip() if block.group(11) else None,
                        "inner_path": block.group(12).strip() if block.group(12) else None,
                        "inner_update_type": block.group(13).strip() if block.group(13) else None,
                        "sql_command": block.group(14).strip() if block.group(14) else None,
                        "sheet": block.group(15).strip() if block.group(15) else None,
                        "cell": block.group(16).strip() if block.group(16) else None,
                        "excel_new_value": block.group(17).strip() if block.group(17) else None,
                        "metadata": block.group(18).strip() if block.group(18) else None,
                        "pattern": block.group(19).strip() if block.group(19) else None,
                        "replacement": block.group(20).strip() if block.group(20) else None,
                        "description": block.group(21).strip() if block.group(21) else None,
                        "content_type": block.group(22).strip(),
                        "if_not_exists": block.group(23).strip(),
                        "directory": block.group(24).strip() if block.group(24) else None,
                        "content": block.group(25).strip()
                    }
                    updates["code"].append(update_entry)
            elif title == "Changelog Updates" and "No updates" not in body:
                entry = re.search(r"- \*\*Entry\*\*:\n---\n(.*?)---", body, re.DOTALL)
                if entry:
                    updates["changelog"] = entry.group(1).strip()
            elif title == "Memory Log Updates" and "No updates" not in body:
                entry = re.search(r"- \*\*Entry\*\*:\n---\n(.*?)---", body, re.DOTALL)
                if entry:
                    updates["memory_log"] = entry.group(1).strip()
            elif title == "Action Items":
                updates["action_items"] = [item.strip() for item in re.findall(r"- (.*?)\n", body)]
        return updates

    def validate_files(self) -> bool:
        try:
            with open(self.project_log_file, "r", encoding="utf-8") as f:
                content = f.read()
            memory_log = content.split("## Memory Log")[1].split("## Memory Log Backup")[0]
            memory_log_backup = content.split("## Memory Log Backup")[1].split("## Changelog")[0]
            if memory_log != memory_log_backup:
                self.logger.error("Memory log and backup mismatch in PROJECT_LOG.md")
                return False
            return True
        except Exception as e:
            self.logger.error(f"Failed to validate files: {e}", exc_info=True)
            return False

    def resolve_filepath(self, relative_path: str, directory: Optional[str] = None) -> str:
        base_path = os.path.join(self.project_dir, directory or "")
        filepath = os.path.abspath(os.path.join(base_path, relative_path))
        if not filepath.startswith(self.project_dir):
            self.logger.error(f"Path {filepath} outside project directory (potential path traversal)")
            raise ValueError(f"Path {filepath} outside project directory")
        return filepath

    def generate_context_summary(self, updates: Dict[str, Any], update_results: List[Dict[str, Any]], status: str) -> None:
        file_structure = "\n".join([f"- {os.path.relpath(os.path.join(root, file), self.project_dir)}" for root, _, files in os.walk(self.project_dir) for file in files if not file.endswith(".backup")])
        updated_files = "\n".join([f"- {result['filepath']} (Version {result['new_version']})" + (f"\n  - Diff: ```diff\n{result['diff']}\n```" if result.get('diff') else "") for result in update_results])
        context_summary = f"""# Context Summary

## Metadata
- **Timestamp**: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
- **Previous Message Reference**: {updates["message_ref"]}
- **Previous Session Summary Version**: {updates["version"]}

## Status
- **Result**: {status}
- **Updated Files**:
{updated_files or "- None"}

## Previous Session Summary
{updates["memory_log"] or "No updates"}

## File Structure
{file_structure}

## Files to Provide
- {self.project_log_file}
{''.join([f'- {result["filepath"]}\n' for result in update_results])}

## Action Items
{''.join([f'- {item}\n' for item in updates["action_items"]]) or "- None"}
"""
        self.io_handler.write_context_summary(self.context_summary_file, context_summary)
        self.logger.info("Generated context summary")

    def process_instruction(self, instruction: str) -> bool:
        self.logger.info(f"Received instruction: {instruction}")
        instruction = instruction.lower()

        def execute_instruction():
            if instruction == "update session":
                return self.process_session_summary()
            elif instruction == "validate session":
                return self.validate_only()
            elif instruction.startswith("rollback"):
                parts = instruction.split()
                file = parts[1] if len(parts) > 1 else None
                return self.rollback(file)
            elif instruction == "list versions":
                self.list_versions()
                return True
            elif instruction == "list locks":
                self.list_locks()
                return True
            elif instruction == "reset versions":
                self.reset_versions()
                return True
            elif instruction == "reset locks":
                self.reset_locks()
                return True
            else:
                self.logger.error(f"Unsupported instruction: {instruction}")
                return False

        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(execute_instruction)
            try:
                return future.result(timeout=self.config["timeout"])
            except FutureTimeoutError:
                self.logger.error(f"Instruction '{instruction}' timed out after {self.config['timeout']} seconds")
                return False

    def validate_only(self) -> bool:
        if not os.path.exists(self.session_file):
            self.logger.error(f"Session summary not found at {self.session_file}")
            return False
        content = self.io_handler.read_session_summary(self.session_file)
        is_valid, errors = self.validate_session_summary(content)
        if not is_valid:
            for error in errors:
                self.logger.error(error)
        return is_valid

    def rollback(self, file: Optional[str] = None) -> bool:
        if not os.path.exists(self.rollback_file):
            self.logger.error("No rollback information available")
            return False
        with open(self.rollback_file, "r") as f:
            rollback_info = json.load(f)
        if file:
            for entry in rollback_info["session"]:
                if os.path.basename(entry["filepath"]) == os.path.basename(file):
                    self.restore_file(entry["filepath"], entry["backup_path"])
                    rollback_info["session"] = [e for e in rollback_info["session"] if os.path.basename(e["filepath"]) != os.path.basename(file)]
                    with open(self.rollback_file, "w") as f:
                        json.dump(rollback_info, f, indent=4)
                    return True
            self.logger.error(f"No rollback info for {file}")
            return False
        else:
            with ThreadPoolExecutor(max_workers=self.config["parallel_workers"]) as executor:
                futures = [executor.submit(self.restore_file, entry["filepath"], entry["backup_path"]) for entry in rollback_info["session"]]
                for future in futures:
                    try:
                        future.result(timeout=self.config["timeout"])
                    except FutureTimeoutError:
                        self.logger.error(f"Rollback timed out after {self.config['timeout']} seconds")
            rollback_info["session"] = []
            with open(self.rollback_file, "w") as f:
                json.dump(rollback_info, f, indent=4)
            return True

    def list_versions(self) -> None:
        versions = self.load_versions()
        self.io_handler.report_status("File Versions:", "info")
        for filename, version in versions.items():
            self.io_handler.report_status(f"- {filename}: {version}", "info")

    def list_locks(self) -> None:
        locks = self.load_locks()
        self.io_handler.report_status("File Locks:", "info")
        if not locks:
            self.io_handler.report_status("- None", "info")
        for filename, timestamp in locks.items():
            self.io_handler.report_status(f"- {filename}: {timestamp}", "info")

    def reset_versions(self) -> None:
        with open(self.versions_file, "w") as f:
            json.dump({}, f)
        self.logger.info("Reset all file versions")

    def reset_locks(self) -> None:
        with self.file_lock_context(self.locks_file):
            with open(self.locks_file, "w") as f:
                json.dump({}, f)
        self.logger.info("Reset all file locks")

def parse_arguments():
    parser = argparse.ArgumentParser(description="Context Manager for AI-driven file editing")
    parser.add_argument("project_dir", nargs="?", default=os.path.dirname(os.path.abspath(__file__)),
                        help="Project directory (default: script directory)")
    parser.add_argument("--verbose", action="store_true", help="Enable verbose output")
    parser.add_argument("--debug", action="store_true", help="Enable debug logging to console")
    parser.add_argument("--gui", action="store_true", help="Run in GUI mode")
    parser.add_argument("--api", action="store_true", help="Run in API mode")
    parser.add_argument("--api-host", default="localhost", help="API host (default: localhost)")
    parser.add_argument("--api-port", type=int, default=8000, help="API port (default: 8000)")
    parser.add_argument("--event", action="store_true", help="Run in event-driven mode")
    parser.add_argument("--list-versions", action="store_true", help="List all file versions")
    parser.add_argument("--list-locks", action="store_true", help="List all file locks")
    parser.add_argument("--reset-versions", action="store_true", help="Reset all file versions")
    parser.add_argument("--reset-locks", action="store_true", help="Reset all file locks")
    parser.add_argument("--validate-only", action="store_true", help="Validate session summary without applying updates")
    parser.add_argument("--rollback", nargs="?", const=True, metavar="FILE", help="Rollback all updates or a specific file")
    parser.add_argument("--profile", action="store_true", help="Enable performance profiling")
    parser.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT, help=f"Timeout in seconds (default: {DEFAULT_TIMEOUT})")
    parser.add_argument("--retry-attempts", type=int, default=3, help="Number of retry attempts (default: 3)")
    parser.add_argument("--retry-delay", type=float, default=1.0, help="Delay between retries in seconds (default: 1.0)")
    return parser.parse_args()

if __name__ == "__main__":
    args = parse_arguments()
    logging.basicConfig(level=logging.DEBUG if args.debug else logging.INFO)
    if args.debug:
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
        logging.getLogger().addHandler(console_handler)

    config = {
        "profile_mode": args.profile,
        "timeout": args.timeout,
        "retry_attempts": args.retry_attempts,
        "retry_delay": args.retry_delay
    }

    if args.gui:
        try:
            from tkinter import Tk, messagebox
            root = Tk()
            root.withdraw()
            io_handler = CommandLineIOHandler(verbose=args.verbose)
            manager = ContextManager(args.project_dir, io_handler, config)
            if manager.process_session_summary():
                messagebox.showinfo("Context Manager", "Session processed successfully")
            else:
                messagebox.showerror("Context Manager", "Failed to process session")
            root.destroy()
            sys.exit(0)
        except ImportError:
            logging.error("GUI mode requires tkinter. Install it with 'pip install tk'")
            sys.exit(1)
        except Exception as e:
            logging.error(f"GUI mode failed: {e}", exc_info=True)
            messagebox.showerror("Context Manager", f"Error: {e}")
            sys.exit(1)

    elif args.api:
        io_handler = APIIOHandler(logging.getLogger(__name__), host=args.api_host, port=args.api_port)
        manager = ContextManager(args.project_dir, io_handler, config)
        io_handler.start_server()
        logging.info("API server running. Press Ctrl+C to stop.")
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            logging.info("API server stopped")
            sys.exit(0)

    elif args.event:
        try:
            from serenity_events import SerenityEvents
            events = SerenityEvents()
            io_handler = EventIOHandler(events, logging.getLogger(__name__))
            manager = ContextManager(args.project_dir, io_handler, config)
            logging.info("Running in event mode. Waiting for instructions...")
            events.run()
        except ImportError:
            logging.error("Event mode requires serenity_events module")
            sys.exit(1)

    else:
        io_handler = CommandLineIOHandler(verbose=args.verbose)
        manager = ContextManager(args.project_dir, io_handler, config)
        
        if args.list_versions:
            manager.list_versions()
            sys.exit(0)
        elif args.list_locks:
            manager.list_locks()
            sys.exit(0)
        elif args.reset_versions:
            manager.reset_versions()
            sys.exit(0)
        elif args.reset_locks:
            manager.reset_locks()
            sys.exit(0)
        elif args.validate_only:
            success = manager.validate_only()
            sys.exit(0 if success else 1)
        elif args.rollback is not None:
            success = manager.rollback(args.rollback if args.rollback is not True else None)
            sys.exit(0 if success else 1)
        else:
            success = manager.process_session_summary()
            sys.exit(0 if success else 1)